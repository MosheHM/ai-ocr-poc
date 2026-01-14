
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict
from ..schemas.base import ValidationResult, ValidationStatus

class ExcelReportBuilder:
    """Creates Excel reports with validation highlighting."""
    
    COLORS = {
        'error_bg': 'FF6B6B',       # Red for failures
        'warning_bg': 'FFE66D',     # Yellow for warnings
        'header_bg': '2C3E50',      # Dark header
        'header_text': 'FFFFFF',    # White text
    }

    def __init__(self):
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            
    def build_report(
        self,
        validated_dfs: Dict[str, pd.DataFrame],
        validation_errors: List[ValidationResult],
        output_path: Path,
    ) -> Path:
        """
        Build complete Excel report with separate sheets per document type.
        
        Args:
            validated_dfs: Dictionary of {DocType: DataFrame}
            validation_errors: List of all validation errors
            output_path: Target path for the .xlsx file
        """
        
        for doc_type, df in validated_dfs.items():
            if df.empty:
                continue
                
            sheet_errors = [e for e in validation_errors if e.doc_type == doc_type]
            
            self._create_sheet(doc_type, df, sheet_errors)

        self._add_validation_summary(validation_errors)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        return output_path

    def _create_sheet(self, sheet_name: str, df: pd.DataFrame, sheet_errors: List[ValidationResult]):
        clean_name = sheet_name.replace("/", "-")[:30]
        ws = self.wb.create_sheet(clean_name)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_header_styles(ws, len(df.columns))
        
        self._highlight_errors(ws, sheet_errors, list(df.columns))
        
        self._auto_fit_columns(ws)

    def _apply_header_styles(self, ws, col_count):
        header_fill = PatternFill(start_color=self.COLORS['header_bg'], end_color=self.COLORS['header_bg'], fill_type='solid')
        header_font = Font(bold=True, color=self.COLORS['header_text'])
        
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    def _highlight_errors(self, ws, errors: List[ValidationResult], columns: List[str]):
        error_fill = PatternFill(start_color=self.COLORS['error_bg'], end_color=self.COLORS['error_bg'], fill_type='solid')
        
        for error in errors:
            if error.column in columns:
                col_idx = columns.index(error.column) + 1
                row_idx = error.row + 2 # +1 header, +1 0-based index
                
                if row_idx <= ws.max_row and col_idx <= ws.max_column:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = error_fill

    def _add_validation_summary(self, errors: List[ValidationResult]):
        if "Sheet" in self.wb.sheetnames:
            ws = self.wb["Sheet"]
            ws.title = "Validation Summary"
        else:
            ws = self.wb.create_sheet("Validation Summary", 0) 
            
        ws['A1'] = "Validation Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Total Errors:"
        ws['B3'] = len([e for e in errors if e.status == ValidationStatus.FAILED])
        
        if errors:
            ws['A5'] = "Detailed Errors"
            ws['A5'].font = Font(bold=True)
            headers = ['Row', 'Column', 'Status', 'Message', 'Original Value']
            for col, h in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=h).font = Font(bold=True)
            
            for i, err in enumerate(errors, 7):
                ws.cell(row=i, column=1, value=err.row)
                ws.cell(row=i, column=2, value=err.column)
                ws.cell(row=i, column=3, value=err.status.value)
                ws.cell(row=i, column=4, value=err.message)
                ws.cell(row=i, column=5, value=str(err.original_value))
                
        self._auto_fit_columns(ws)

    def _auto_fit_columns(self, ws):
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
